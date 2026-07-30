
# ============================================================
#  DAY-AHEAD MULTI-MARKET FORECASTER  
#
#  9am run (no flag)  : forecast tomorrow (DAM/GDAM/RTM, 96 blocks each),
#                       write colour-coded Excel, backfill any actuals.
#  --backfill-only    : only fill in actuals + refresh accuracy (no re-forecast).
#
#  v2 changes:
#    * cap-regime features (capcount/evecap/capstreak/blkcapfreq7) per market
#    * cap classifier + gate for ALL THREE markets (was RTM only)
#    * rank-hit uses a 5% tolerance (a pick counts if within 5% of the true best)
#    * strict rank-hit kept alongside for comparison
#    * a "Scored" Excel sheet colours each past block green(hit)/red(miss)
# ============================================================
#str(OUT_DIR) str(OUT_DIR)
#str(OUT_DIR) str(OUT_DIR) --backfill-only
import os
import sys
import json
import datetime
import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import xgboost as xgb
import requests
import holidays as hl

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------
#  PATHS
# ------------------------------------------------------------
from .config import MODEL_DIR, OUT_DIR, IEX_TOKEN, RANK_TOL
LOG_FILE  = os.path.join(OUT_DIR, "dayahead_forecast_log.csv")
HEARTBEAT = os.path.join(OUT_DIR, "dayahead_run_log.txt")


def heartbeat(msg):
    stamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    line = f"[{stamp}] {msg}"
    print(line)
    try:
        with open(HEARTBEAT, "a") as f:
            f.write(line + "\n")
    except Exception:
        pass

# ------------------------------------------------------------
#  1. LOAD MODELS + CONFIG
# ------------------------------------------------------------
# Loaded lazily so that the data fetchers below can be imported (e.g. by the
# retraining pipeline) before any model artefacts exist on disk.
CAP_LVL = 9900
CFG = None
MARKETS = FEATURES = ANCHORS = BASE_SHIFT = None
CAL_FEATS = WX_FEATS = CAP_THRESHOLD = GATE = CAP_VALUE = None
REG, CLF = {}, {}


def load_models():
    """Read dayahead_config.json and the six XGBoost artefacts from MODEL_DIR."""
    global CFG, MARKETS, FEATURES, ANCHORS, BASE_SHIFT
    global CAL_FEATS, WX_FEATS, CAP_THRESHOLD, GATE, CAP_VALUE

    cfg_path = os.path.join(MODEL_DIR, "dayahead_config.json")
    if not os.path.exists(cfg_path):
        raise FileNotFoundError(
            f"No model config at {cfg_path}. Run the retraining pipeline first: "
            "python -m iex_dayahead.retrain"
        )
    with open(cfg_path) as f:
        CFG = json.load(f)

    MARKETS       = CFG["markets"]
    FEATURES      = CFG["features"]
    ANCHORS       = CFG["anchors"]
    BASE_SHIFT    = CFG["base_shift"]
    CAL_FEATS     = CFG["calendar_feats"]
    WX_FEATS      = CFG["weather_feats"]
    CAP_THRESHOLD = CFG["cap_threshold"]
    GATE          = CFG["gate"]
    CAP_VALUE     = CFG["cap_value"]

    for m in MARKETS:
        r = xgb.XGBRegressor()
        r.load_model(os.path.join(MODEL_DIR, f"{m}_regressor.json"))
        REG[m] = r
        c = xgb.XGBClassifier()
        c.load_model(os.path.join(MODEL_DIR, f"{m}_classifier.json"))
        CLF[m] = c
    return CFG

# ------------------------------------------------------------
#  2. LIVE DATA FETCHERS  (IEX API)
# ------------------------------------------------------------
IEX_URL   = "https://www.iexindia.com/IEXPublish/AppServices.svc/IEXGetTradeData"

def _require_token():
    if not IEX_TOKEN:
        raise RuntimeError(
            "IEX_TOKEN environment variable not set. "
            "Request an API token from IEX and export it before running."
        )
    return IEX_TOKEN
PCODE = {"dam": 1, "gdam": 2, "rtm": 3}

def fetch_iex(market, from_date, to_date):
    payload = json.dumps({"APITokenNo": _require_token(), "Product_Code": PCODE[market],
        "From_Date": from_date, "From_Token": 1, "To_Date": to_date, "To_Token": 96, "Date_Type": 2})
    r = requests.post(IEX_URL, headers={'Content-Type': 'application/json'}, data=payload, timeout=120)
    r.raise_for_status()
    rows = []
    for day in r.json().get("Delivery_Date_Details", []):
        dt = pd.to_datetime(day["DeliveryDate"], dayfirst=True).normalize()
        for i, tok in enumerate(day["Token_Wise"]):
            a = tok["All_India_DAM_GDAM_RTM"]
            rows.append({"date": dt, "block": i, "mcp": a.get("Clearing_Price"),
                         "pb": a.get("Buy_Volume"), "sb": a.get("Sell_Volume"),
                         "mcv": a.get("Cleared_Volume")})
    return pd.DataFrame(rows)

FC_URL = "https://api.open-meteo.com/v1/forecast"
HVARS  = ["shortwave_radiation", "temperature_2m", "cloud_cover", "wind_speed_100m"]
loc_solar = {"bhadla":(27.53,71.91),"pavagada":(14.10,77.27),"kurnool":(15.82,78.03),"charanka":(23.90,71.20)}
loc_wind  = {"muppandal":(8.25,77.53),"jaisalmer":(26.91,70.90),"brahmanvel":(21.15,74.31),"kutch":(23.73,69.85)}
loc_load  = {"delhi":(28.61,77.21),"mumbai":(19.07,72.87),"chennai":(13.08,80.27),"bangalore":(12.97,77.59)}

def _fc(lat, lon):
    r = requests.get(FC_URL, params={"latitude":lat,"longitude":lon,"hourly":",".join(HVARS),
                     "timezone":"Asia/Kolkata","forecast_days":3,"past_days":2}, timeout=60)
    r.raise_for_status()
    dd = pd.DataFrame(r.json()["hourly"]); dd["time"] = pd.to_datetime(dd["time"])
    return dd.set_index("time")[HVARS]

def _avg(locs):
    return sum(_fc(la, lo) for la, lo in locs.values()) / len(locs)

def fetch_weather_forecast():
    solar, wind, load = _avg(loc_solar), _avg(loc_wind), _avg(loc_load)
    wx = pd.DataFrame(index=solar.index)
    wx["ghi"]  = solar["shortwave_radiation"]; wx["wind100"] = wind["wind_speed_100m"]
    wx["temp"] = load["temperature_2m"];       wx["cloud"]   = (solar["cloud_cover"] + load["cloud_cover"]) / 2
    wx = wx.reset_index(); wx["date"] = wx["time"].dt.normalize(); wx["hour"] = wx["time"].dt.hour
    return wx

# ------------------------------------------------------------
#  3. BUILD TOMORROW'S FEATURE TABLE  (mirrors training notebook)
# ------------------------------------------------------------
def _grid_with_target(d, col, target_day):
    g = d.pivot(index="date", columns="block", values=col).sort_index()
    if target_day not in g.index:
        g.loc[target_day] = np.nan
        g = g.sort_index()
    return g

def build_tomorrow_features(target_day):
    frm = (target_day - pd.Timedelta(days=25)).strftime('%d/%m/%Y')
    to  = target_day.strftime('%d/%m/%Y')

    raw = {}
    for m in MARKETS:
        f = fetch_iex(m, frm, to)
        f = f.rename(columns={"mcp": m, "pb": f"{m}_pb", "sb": f"{m}_sb", "mcv": f"{m}_mcv"})
        raw[m] = f[["date","block",m,f"{m}_pb",f"{m}_sb",f"{m}_mcv"]]
    d = raw["dam"].merge(raw["gdam"], on=["date","block"], how="outer") \
                  .merge(raw["rtm"], on=["date","block"], how="outer") \
                  .sort_values(["date","block"]).reset_index(drop=True)

    grids = {m: _grid_with_target(d, m, target_day) for m in MARKETS}

    # price-history features (market-correct shift)
    price_feats = {}
    for m in MARKETS:
        P = grids[m]; s = BASE_SHIFT[m]
        price_feats[f"{m}_lag_recent"] = P.shift(s)
        price_feats[f"{m}_lag_prev"]   = P.shift(s+1)
        price_feats[f"{m}_lag7d"]      = P.shift(7)
        price_feats[f"{m}_lag14d"]     = P.shift(14)
        price_feats[f"{m}_roll7_mean"]  = P.shift(s).rolling(7).mean()
        price_feats[f"{m}_roll7_std"]   = P.shift(s).rolling(7).std()
        price_feats[f"{m}_roll14_mean"] = P.shift(s).rolling(14).mean()

    L = pd.DataFrame({"date": target_day, "block": range(96)})
    ts = pd.Timestamp(target_day)
    L["hour"]    = L["block"] // 4
    L["dow"]     = ts.dayofweek
    L["is_wknd"] = int(ts.dayofweek >= 5)
    L["month"]   = ts.month
    L["doy"]     = ts.dayofyear
    L["sin_blk"] = np.sin(2*np.pi*L["block"]/96); L["cos_blk"] = np.cos(2*np.pi*L["block"]/96)
    L["sin_doy"] = np.sin(2*np.pi*L["doy"]/365);  L["cos_doy"] = np.cos(2*np.pi*L["doy"]/365)
    in_hol = pd.to_datetime(list(hl.India(years=range(2023,2029)).keys()))
    L["is_hol"]  = int(ts.normalize() in set(in_hol))

    def map_row(grid):
        return L["block"].map(grid.loc[target_day].to_dict()) if target_day in grid.index else np.nan

    for m in MARKETS:
        s = grids[m].shift(BASE_SHIFT[m])
        def sm(func):
            return func(s.loc[ts]) if ts in s.index else np.nan
        L[f"{m}_daymean"] = sm(lambda r: r.mean())
        L[f"{m}_daymax"]  = sm(lambda r: r.max())
        L[f"{m}_daymin"]  = sm(lambda r: r.min())
        L[f"{m}_evening"] = sm(lambda r: r[list(range(68,88))].mean())
        L[f"{m}_midday"]  = sm(lambda r: r[list(range(44,56))].mean())

    for name, grid in price_feats.items():
        L[name] = map_row(grid)

    for m in MARKETS:
        s = BASE_SHIFT[m]
        for side in ["pb","sb","mcv"]:
            g = _grid_with_target(d, f"{m}_{side}", target_day).shift(s)
            L[f"{m}_{side}_lag"] = map_row(g)
        gpb = _grid_with_target(d, f"{m}_pb", target_day).shift(s)
        gsb = _grid_with_target(d, f"{m}_sb", target_day).shift(s)
        if target_day in gpb.index:
            L[f"{m}_sdgap_lag"] = L["block"].map((gpb.loc[target_day]-gsb.loc[target_day]).to_dict())
        else:
            L[f"{m}_sdgap_lag"] = np.nan
    L["spread_dam_gdam"] = L["dam_lag_recent"] - L["gdam_lag_recent"]
    L["spread_dam_rtm"]  = L["dam_lag_recent"] - L["rtm_lag_recent"]

    # cap-regime features (must match notebook Cell 3C exactly)
    for m in MARKETS:
        P = grids[m]; s = BASE_SHIFT[m]
        capped = (P >= CAP_LVL).astype(float)
        day_cap_count = capped.sum(axis=1)
        eve_capped    = (capped[list(range(68,88))].sum(axis=1) > 0).astype(int)
        streak = (day_cap_count > 0).astype(int)
        streak = streak.groupby((streak == 0).cumsum()).cumsum()
        cc = day_cap_count.shift(s); ec = eve_capped.shift(s); st = streak.shift(s)
        L[f"{m}_capcount_recent"] = cc.loc[ts] if ts in cc.index else np.nan
        L[f"{m}_evecap_recent"]   = ec.loc[ts] if ts in ec.index else np.nan
        L[f"{m}_capstreak"]       = st.loc[ts] if ts in st.index else np.nan
        blk_freq = capped.shift(s).rolling(7).mean()
        L[f"{m}_blkcapfreq7"] = map_row(blk_freq)

    try:
        wx = fetch_weather_forecast()
        wxd = wx[wx["date"] == ts]
        hm = wxd.set_index("hour")
        for c in WX_FEATS:
            L[c] = L["hour"].map(hm[c].to_dict()) if len(wxd) else np.nan
    except Exception as e:
        heartbeat(f"weather fetch failed ({e}); proceeding without weather")
        for c in WX_FEATS:
            L[c] = np.nan

    return L

# ------------------------------------------------------------
#  4. FORECAST all three markets for tomorrow (with cap gate)
# ------------------------------------------------------------
def forecast_day(target_day):
    L = build_tomorrow_features(target_day)
    out = pd.DataFrame({"date": target_day, "block": range(96)})
    out["time"] = [f"{b*15//60:02d}:{b*15%60:02d}" for b in range(96)]

    for m in MARKETS:
        feats = FEATURES[m]
        for c in feats:
            if c not in L.columns:
                L[c] = np.nan
        X = L[feats].values
        anchor = L[ANCHORS[m]].values
        yhat = anchor + REG[m].predict(X)
        cap_p = CLF[m].predict_proba(X)[:, 1]
        yhat = np.where(cap_p >= GATE, (1-cap_p)*yhat + cap_p*CAP_VALUE, yhat)
        yhat = np.clip(yhat, 0, 12000)
        out[m] = pd.Series(np.round(yhat), index=out.index).astype("Int64")
        out[f"{m}_cap_prob"] = np.round(cap_p, 2)

    price_cols = out[MARKETS]
    out["cheapest"] = price_cols.idxmin(axis=1).str.upper()
    out["dearest"]  = price_cols.idxmax(axis=1).str.upper()
    return out

# ------------------------------------------------------------
#  5. WRITE COLOUR-CODED EXCEL
# ------------------------------------------------------------
GREEN = PatternFill("solid", fgColor="C6EFCE")
RED   = PatternFill("solid", fgColor="FFC7CE")
HDR   = PatternFill("solid", fgColor="1F3864")
THIN  = Side(style="thin", color="D9D9D9"); BORDER = Border(THIN,THIN,THIN,THIN)

def _style_header(ws, ncol):
    for col in range(1, ncol+1):
        c = ws.cell(row=1, column=col)
        c.fill = HDR; c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.alignment = Alignment(horizontal="center"); c.border = BORDER

def write_excel(fc, target_day, log=None):
    path = os.path.join(OUT_DIR, f"dayahead_forecast_{target_day.strftime('%Y%m%d')}.xlsx")
    wb = Workbook(); ws = wb.active; ws.title = "Forecast"

    headers = ["Block","Time","DAM (Rs)","GDAM (Rs)","RTM (Rs)","Cheapest","Dearest"]
    ws.append(headers); _style_header(ws, len(headers))
    mkt_col = {"DAM":3, "GDAM":4, "RTM":5}
    for _, r in fc.iterrows():
        ws.append([int(r["block"])+1, r["time"],
                   int(r["dam"]) if pd.notna(r["dam"]) else None,
                   int(r["gdam"]) if pd.notna(r["gdam"]) else None,
                   int(r["rtm"]) if pd.notna(r["rtm"]) else None,
                   r["cheapest"], r["dearest"]])
        er = ws.max_row
        for col in range(1, len(headers)+1):
            cc = ws.cell(row=er, column=col)
            cc.font = Font(name="Arial", size=10); cc.alignment = Alignment(horizontal="center"); cc.border = BORDER
        if r["dearest"] in mkt_col:  ws.cell(row=er, column=mkt_col[r["dearest"]]).fill = GREEN
        if r["cheapest"] in mkt_col: ws.cell(row=er, column=mkt_col[r["cheapest"]]).fill = RED
    for i, w in enumerate([7,8,11,12,11,11,11], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ---- Scored sheet: most recent fully-scored day, coloured by hit/miss ----
    if log is not None:
        scored = log.dropna(subset=["dam_act","gdam_act","rtm_act"])
        if len(scored):
            day = scored["target_date"].max()
            dd = log[log["target_date"] == day].sort_values("block")
            ws2 = wb.create_sheet("Scored")
            h2 = ["Block","Time","DAM fc","DAM act","GDAM fc","GDAM act","RTM fc","RTM act",
                  "Pred dearest","Actual dearest","Hit(5%)","Hit(strict)"]
            ws2.append(h2); _style_header(ws2, len(h2))
            for _, r in dd.iterrows():
                ws2.append([int(r["block"])+1, r["time"],
                            r.get("dam_fc"), r.get("dam_act"),
                            r.get("gdam_fc"), r.get("gdam_act"),
                            r.get("rtm_fc"), r.get("rtm_act"),
                            r.get("max_pred"), r.get("max_actual"),
                            r.get("max_hit"), r.get("max_hit_strict")])
                er = ws2.max_row
                for col in range(1, len(h2)+1):
                    cc = ws2.cell(row=er, column=col)
                    cc.font = Font(name="Arial", size=10); cc.alignment = Alignment(horizontal="center"); cc.border = BORDER
                hitcell = ws2.cell(row=er, column=11)
                if r.get("max_hit") == "Y": hitcell.fill = GREEN
                elif r.get("max_hit") == "N": hitcell.fill = RED
            for i, w in enumerate([7,8,9,9,9,9,9,9,13,14,9,11], start=1):
                ws2.column_dimensions[get_column_letter(i)].width = w
            ws2.freeze_panes = "A2"

            wsL = wb.create_sheet("Legend")
            wsL["A1"] = f"Forecast sheet: tomorrow ({target_day.date()}). Green=dearest market, Red=cheapest market."
            wsL["A3"] = f"Scored sheet: {pd.Timestamp(day).date()} (most recent fully-cleared day)."
            wsL["A4"] = "Hit(5%): model's dearest-market pick was within 5% of the true dearest price."
            wsL["A5"] = "Green = hit, Red = miss."
            wsL["A5"].fill = GREEN
            for r_ in [1,3,4,5]:
                wsL[f"A{r_}"].font = Font(name="Arial", size=10)
            wsL.column_dimensions["A"].width = 90

    wb.save(path)
    return path

# ------------------------------------------------------------
#  6. RANKING COLUMNS (5% tolerance + strict)
# ------------------------------------------------------------
LOG_COLS = ["target_date","block","time",
            "dam_fc","gdam_fc","rtm_fc",
            "dam_act","gdam_act","rtm_act",
            "max_pred","max_actual","min_pred","min_actual",
            "max_hit","min_hit","max_hit_strict","min_hit_strict"]

ACT_OF   = {"DAM":"dam_act","GDAM":"gdam_act","RTM":"rtm_act"}
NAME_FC  = {"dam_fc":"DAM","gdam_fc":"GDAM","rtm_fc":"RTM"}
NAME_ACT = {"dam_act":"DAM","gdam_act":"GDAM","rtm_act":"RTM"}

def add_ranking_cols(log):
    fc_cols  = ["dam_fc","gdam_fc","rtm_fc"]
    act_cols = ["dam_act","gdam_act","rtm_act"]
    f = log[fc_cols].astype(float)
    log["max_pred"] = f.idxmax(axis=1).map(NAME_FC)
    log["min_pred"] = f.idxmin(axis=1).map(NAME_FC)

    a = log[act_cols].astype(float)
    complete = a.notna().all(axis=1)
    for col in ["max_actual","min_actual","max_hit","min_hit","max_hit_strict","min_hit_strict"]:
        log[col] = pd.Series([np.nan]*len(log), index=log.index, dtype=object)

    if complete.any():
        ac = a[complete]
        log.loc[complete, "max_actual"] = ac.idxmax(axis=1).map(NAME_ACT)
        log.loc[complete, "min_actual"] = ac.idxmin(axis=1).map(NAME_ACT)
        log.loc[complete, "max_hit_strict"] = np.where(
            log.loc[complete,"max_pred"].values==log.loc[complete,"max_actual"].values, "Y","N")
        log.loc[complete, "min_hit_strict"] = np.where(
            log.loc[complete,"min_pred"].values==log.loc[complete,"min_actual"].values, "Y","N")
        for i in log.index[complete]:
            best_max = ac.loc[i].max(); best_min = ac.loc[i].min()
            picked_max_act = log.at[i, ACT_OF[log.at[i,"max_pred"]]]
            picked_min_act = log.at[i, ACT_OF[log.at[i,"min_pred"]]]
            log.at[i,"max_hit"] = "Y" if picked_max_act >= (1-RANK_TOL)*best_max else "N"
            log.at[i,"min_hit"] = "Y" if picked_min_act <= (1+RANK_TOL)*best_min else "N"
    return log

def accuracy_line(log):
    parts = []
    for m in MARKETS:
        s = log.dropna(subset=[f"{m}_act"])
        if len(s):
            w = 100*(s[f"{m}_fc"]-s[f"{m}_act"]).abs().sum()/s[f"{m}_act"].abs().sum()
            parts.append(f"{m.upper()} {100-w:.0f}%({len(s)})")
    scored = log[log["max_hit"].isin(["Y","N"])]
    if len(scored):
        h5 = (scored["max_hit"]=="Y").mean()*100
        hs = (scored["max_hit_strict"]=="Y").mean()*100
        parts.append(f"rank-hit(dearest) 5%:{h5:.0f}% strict:{hs:.0f}% ({len(scored)})")
    return " | ".join(parts) if parts else "no actuals yet"

# ------------------------------------------------------------
#  7. BACKFILL / LOG / PLOT
# ------------------------------------------------------------
def _fetch_actuals_map(m):
    today = datetime.date.today()
    frm = (today - datetime.timedelta(days=4)).strftime('%d/%m/%Y')
    to  = (today + datetime.timedelta(days=1)).strftime('%d/%m/%Y')   # include tomorrow
    a = fetch_iex(m, frm, to)
    a["target_date"] = a["date"]
    return a.set_index(["target_date","block"])["mcp"].to_dict()

def do_backfill(log):
    for m in MARKETS:
        need = log[log[f"{m}_act"].isna()]
        if not len(need):
            continue
        try:
            amap = _fetch_actuals_map(m)
        except Exception as e:
            heartbeat(f"{m} actuals fetch failed ({e})"); continue
        for i in need.index:
            key = (pd.Timestamp(log.at[i,"target_date"]), int(log.at[i,"block"]))
            v = amap.get(key)
            if v is not None and pd.notna(v) and v > 0:
                log.at[i, f"{m}_act"] = v
    return log

def load_log():
    if os.path.exists(LOG_FILE):
        log = pd.read_csv(LOG_FILE, parse_dates=["target_date"])
    else:
        log = pd.DataFrame(columns=LOG_COLS)
    for c in LOG_COLS:
        if c not in log.columns:
            log[c] = np.nan
    return log

def plot_recent(log):
    scored = log.dropna(subset=["dam_act"])
    if not len(scored):
        return None
    day = scored["target_date"].max()
    dd = log[log["target_date"] == day].sort_values("block")
    fig, axes = plt.subplots(3, 1, figsize=(11, 9), sharex=True)
    for ax, m in zip(axes, MARKETS):
        ax.plot(dd["block"], dd[f"{m}_fc"], "--", color="#c00000", label="Forecast", linewidth=2)
        if dd[f"{m}_act"].notna().any():
            ax.plot(dd["block"], dd[f"{m}_act"], color="#1f3864", label="Actual", linewidth=2)
        ax.set_title(f"{m.upper()} — {pd.Timestamp(day).date()}")
        ax.set_ylabel("Rs/MWh"); ax.legend(loc="upper left"); ax.grid(alpha=0.3)
    axes[-1].set_xlabel("Block (0-95)")
    plt.tight_layout()
    ppath = os.path.join(OUT_DIR, f"plot_{pd.Timestamp(day).strftime('%Y%m%d')}.png")
    plt.savefig(ppath, dpi=120); plt.close()
    plot_recent_interactive(log)      # also write the interactive HTML
    return ppath

def plot_recent_interactive(log):
    """Interactive Plotly chart: hover shows forecast vs actual per block. Writes an .html."""
    try:
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
    except ImportError:
        heartbeat("plotly not installed - skipping interactive chart (pip install plotly)")
        return None
    scored = log.dropna(subset=["dam_act"])
    if not len(scored):
        return None
    day = scored["target_date"].max()
    dd = log[log["target_date"] == day].sort_values("block").reset_index(drop=True)
    times = dd["time"].tolist() if "time" in dd.columns else [str(b) for b in dd["block"]]

    fig = make_subplots(rows=3, cols=1, shared_xaxes=True,
                        subplot_titles=[m.upper() for m in MARKETS], vertical_spacing=0.07)
    for r, m in enumerate(MARKETS, start=1):
        fig.add_trace(go.Scatter(
            x=dd["block"], y=dd[f"{m}_fc"], name=f"{m.upper()} forecast",
            mode="lines", line=dict(color="#c00000", dash="dash", width=2),
            customdata=times,
            hovertemplate="Block %{x} (%{customdata})<br>Forecast: ₹%{y:,.0f}<extra></extra>"),
            row=r, col=1)
        if dd[f"{m}_act"].notna().any():
            fig.add_trace(go.Scatter(
                x=dd["block"], y=dd[f"{m}_act"], name=f"{m.upper()} actual",
                mode="lines", line=dict(color="#1f3864", width=2),
                customdata=times,
                hovertemplate="Block %{x} (%{customdata})<br>Actual: ₹%{y:,.0f}<extra></extra>"),
                row=r, col=1)
        fig.update_yaxes(title_text="Rs/MWh", row=r, col=1)
    fig.update_xaxes(title_text="Block (0-95)", row=3, col=1)
    fig.update_layout(height=850, width=1000, hovermode="x unified",
                      title=f"Day-Ahead Forecast vs Actual — {pd.Timestamp(day).date()}",
                      template="plotly_white", legend=dict(orientation="h", y=-0.08))
    hpath = os.path.join(OUT_DIR, f"plot_{pd.Timestamp(day).strftime('%Y%m%d')}.html")
    fig.write_html(hpath)
    return hpath

# ------------------------------------------------------------
#  RUN
# ------------------------------------------------------------
def run_forecast():
    target = pd.Timestamp(datetime.date.today() + datetime.timedelta(days=1)).normalize()
    fc = forecast_day(target)

    log = load_log()
    if not (log["target_date"] == pd.Timestamp(target)).any():
        new = pd.DataFrame({
            "target_date": pd.Timestamp(target), "block": fc["block"], "time": fc["time"],
            "dam_fc": fc["dam"], "gdam_fc": fc["gdam"], "rtm_fc": fc["rtm"],
        })
        log = pd.concat([log, new], ignore_index=True)
    log = do_backfill(log)
    log = add_ranking_cols(log)
    log = log[LOG_COLS]
    log.to_csv(LOG_FILE, index=False)

    xlsx = write_excel(fc, target, log)
    plot_recent(log)
    heartbeat(f"forecast {target.date()} written ({os.path.basename(xlsx)}) | {accuracy_line(log)}")

def run_backfill_only():
    if not os.path.exists(LOG_FILE):
        heartbeat("backfill: no log yet"); return
    log = load_log()
    log = do_backfill(log)
    log = add_ranking_cols(log)
    log = log[LOG_COLS]
    log.to_csv(LOG_FILE, index=False)
    plot_recent(log)
    heartbeat(f"backfill run | {accuracy_line(log)}")

if __name__ == "__main__":
    try:
        if "--backfill-only" in sys.argv:
            run_backfill_only()
        else:
            run_forecast()
    except Exception as e:
        heartbeat(f"ERROR: {type(e).__name__}: {e}")
        raise
